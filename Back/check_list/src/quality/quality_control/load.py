import copy
import os
import re
import platform
from openpyxl import load_workbook
from enum import Enum
from openpyxl.cell import MergedCell, Cell
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from quality.inform_for_method.models import Inform_for_method
import xlwings as xw
from quality.quality_control.models import Sample

PETROPHYSICS_VALUE = [0.35, 0.25, 0.20, 0.10, 0.10]
PETROPHYSICS = {}
for i, key in enumerate(Inform_for_method.Types_task):
    PETROPHYSICS.update({key: PETROPHYSICS_VALUE[i]})

DEFAULT = 15
AUTHOR = ""


class SheetNames(Enum):
    """список названий листов"""
    SHEET_1 = "Лист1. Общая информация"
    SHEET_2 = "Лист2. Оценка качества ГИС"
    SHEET_3 = "Лист3. Оценка петрозадачи"
    SHEET_4 = "Лист4. Графика"


"""
петрофизические задачи с распределенными коэффициентами
WIDTH = [27.00, 16.00, 16.00, 16.00, 16.00, 16.00, 23.00, 23.00, 23.00, 23.00, 23.00, 23.00, 11.00, 40.00]
keys_full_inform = ["titul_list_full_inform", "well_construction_full_inform", "wellbore_sizes_full_inform",
                    "chrono_data_full_inform", "sol_data_full_inform", "dash_comp_full_inform",
                    "summary_data_full_inform",
                    "inklino_data_full_inform", "main_record_full_inform", "parametr_full_inform",
                    "control_record_full_inform", "lqc_full_inform", "calibration_full_inform"]
keys_las_file = ["cap_las_file", "parametres_las_file", "mnemodescription_las_file", "tabledata_las_file"]
keys_witsml = ["fullness_data_witsml", "curvenames_witsml", "mnemodescription_witsml"]
"""

"""цвета и соответствующие им значения: string"""
COLOR_STRING = {
    "00B050": ["Имеется", "Увязан", "Отсутствуют", "Отсутствует", "Соответствует основному замеру",
               "Соответствует данным опорных скважин", "Лежат в области ожидаемых значений", "Соответствуют",
               "Хорошее", "Коррелируют"],
    "0000FF": ["Частично неувязан", "Единичные", "Незначительная", "Частично коррелируют",
               "Отменена по согласованию сторон", "Не соответствует основной записи",
               "Не соответствует данным опорных скважин (занижено)", "Завышены",
               "Реперные горизонты не вскрыты", "Свойства реперного горизонта отсутствуют",
               "Не соответствует данным опорных скважин (завышено)", "Занижены", "Удовлетворительное"],
    "FF0000": ["Не имеется", "Неувязан", "Регулярные", "Высокая",
               "Не произведена", "Не лежат в области ожидаемых значений", "Не соответствуют",
               "Свойства реперного горизонта расходятся", "Неудовлетворительное", "Не коррелируют"]}

"""цвета и соответствующие им значения: первый лист (full_inform/digital)"""
COLOR_STRING_GENERAL = {
    "00B050": ["Полная", "Имеется"], "0000FF": "Частичная", "FF0000": ["Отсутствует", "Не имеется"]
}

"""цвета и соответствующие им значения: числа"""
COLOR = {"00B050": [0.85, 1.00], "0000FF": [0.65, 0.84], "FF0000": [0.00, 0.64]}
COLOR_FAIL = {"00B050": [0.00, 0.10], "0000FF": [0.10, 0.20], "FF0000": [0.20, 1.00]}


def style_row(ws, amount_method):
    """определяется стиль для ячеек с методами"""
    for i in range(1, amount_method + 1):
        row = i + 11
        for row in ws.iter_rows(min_row=row, min_col=1, max_col=14, max_row=row):
            for cell in row:
                cell.border = Border(left=Side(border_style="medium", color='FFFFFF'),
                                     right=Side(border_style="medium", color='FFFFFF'),
                                     top=Side(border_style="medium", color='FFFFFF'),
                                     bottom=Side(border_style="medium", color='FFFFFF'))
                cell.fill = PatternFill("solid", fgColor="faf0c8")


def get_petrophysic_string(petrophysic_selected):
    """формируется строка со списком петрофизических задач"""
    petrophysic_string = ""
    amount_petrophysic = len(petrophysic_selected) - 1
    if len(petrophysic_selected) > 0:
        for i in range(amount_petrophysic):
            petrophysic_string = petrophysic_string + f"{petrophysic_selected[i]};\n"
        petrophysic_string = petrophysic_string + f"{petrophysic_selected[amount_petrophysic]}."
    return petrophysic_string


def get_without_spec(data):
    """Формирование нового словаря с данными"""
    data["density"] = data["density"] if isinstance(data["density"], float) else data["density"]["density"]
    without_spec = {}
    without_spec.update(data)
    return without_spec


def not_used_data(data):
    """в ячейки с отсутствующими данными записывается -"""
    if "status" in data:
        if data["status"] is False:
            for key in data.keys():
                data[key] = "-"
    return data


def get_result_description(name, number):
    """находим по задаваемому местоположению причину применения понижающего коэфф"""
    results = [f"На данных ГИС {name} при бурении/из памяти прибора наблюдаются единичные выбросы",
               f"На данных ГИС {name} при бурении/из памяти прибора наблюдаются регулярные выбросы",
               f"На данных ГИС {name} наблюдается незначительная зашумленность",
               f"На данных ГИС {name} наблюдается высокая зашумленность",
               f"Показания метода {name} не соответствуют вскрытому разрезу",
               f"{name} частично коррелирует с другими методами ГИС",
               f"{name} не коррелирует с другими методами ГИС",
               "Отмечается низкая плотность данных реального времени",
               "Во время бурения наблюдалась проблема с передачей данных реального времени",
               "Данные из памяти прибора были перерассчитаны", "Отсутствует Акт промера бурового инструмента",
               "Кривые ГИС частично не увязаны по глубине", "Кривые ГИС не увязаны по глубине",
               "Данные контрольной записи не соответствуют основной записи ГИС",
               "Контрольная запись данных ГИС не произведена",
               "Распределение данных ГИС при бурении на гистограммах не соответствует данным опорных скважин (занижено)",
               "Распределение данных ГИС при бурении на гистограммах не соответствует данным опорных скважин (завышено)",
               "Распределение данных ГИС при бурении на кросс-плотах относительно палеток "
               "не лежит в области ожидаемых значений",
               "Значения абсолютных петрофизических значений в реперных горизонтах занижены",
               "Значения абсолютных петрофизических значений в реперных горизонтах завышены",
               "Реперные горизонты не вскрыты, сравнение абсолютных петрофизических значений "
               "в реперных горизонтах невозможно",
               "Сравнение абсолютных петрофизических значений в реперных горизонтах невозможно, "
               "т.к. свойства реперного горизонта отсуствуют",
               "Сравнение абсолютных петрофизических значений в реперных горизонтах невозможно, "
               "т.к. свойства реперного горизонта расходятся",
               "По показаниям LQC планшетов техническое состояние прибора Удовлетворительное",
               "По показаниям LQC планшетов техническое состояние прибора Неудовлетворительное"]
    return results[number]


def get_value_descr_individ(value_parametr):
    """находим соответствие между значением параметра и списком критических значений для понижающего коэфф,
    возвращаем местоположение в list"""
    values = ["Единичные", "Регулярные", "Незначительная", "Высокая", "Не соответствует", "Частично коррелируют",
              "Не коррелируют", "Низкая плотность данных реального времени",
              "Проблема с передачей данных реального времени", "Перерасчет данных из памяти прибора",
              "Не имеется", "Частично не увязан", "Не увязан",
              "Не соответствует основной записи", "Не произведена",
              "Не соответствует данным опорных скважин (занижено)",
              "Не соответствует данным опорных скважин (завышено)",
              "Не лежат в области ожидаемых значений", "Занижены", "Завышены",
              "Реперные горизонты не вскрыты", "Свойства реперного горизонта отсутствуют",
              "Свойства реперного горизонта расходятся", "Удовлетворительное",
              "Неудовлетворительное"]
    for i in range(len(values)):
        if value_parametr == values[i]:
            return i


def get_description_list(data):
    """формируем список причин по всем методам"""
    second_tables = data["second_table"]
    # part_description = []
    part_description = {}
    comm_names_method = {}
    for method in second_tables:
        for key in method:
            number_result = get_value_descr_individ(method[key])
            if number_result:
                try:
                    name_method = comm_names_method[number_result] + [f'"{method["method"]}"']
                except:
                    name_method = [f'"{method["method"]}"']
                comm_names_method[number_result] = name_method
                part_description[method[key]] = get_result_description(", ".join(set(name_method)), number_result)
    descr_values = list(part_description.values())
    return descr_values, len(descr_values)


def get_descriptions(data):
    """формируем dict с описанием всех причин"""
    description_list, amount_row = get_description_list(data)

    height_descr_list = amount_row * 20 if amount_row != 0 else 60

    description = ""
    if len(description_list) != 0:
        len_descr = len(description_list)
        for i in range(len_descr - 1):
            description = description + f"{i + 1}. {description_list[i]};\n"
        description = description + f"{len_descr}. {description_list[len_descr - 1]}."
    return description, height_descr_list


def special_data(key_sp, data, new_data):
    """формирует данные для full_inform, digital_data"""
    specials = []
    # if key in ["full_inform", "las_file", "witsml"]:
    if key_sp in ["full_inform", "digital_data"]:
        data_special = value_count_status(not_used_data(data[key_sp]))
        # data_special = value_count_status(data_special)
        for key_special in data_special.keys():
            if key_special == "data_status":
                data_special[key_special] = f"Полнота предоставления цифровых данных ({data_special[key_special]})"
            new_data[f"{key_special}"] = data_special[key_special]
            specials.append(f"{key_special}")
        del new_data[key_sp]
    return specials


def petrophysic_selected_num(petrophysic_selected):
    """записывается для каждой петрофизической задачи коэффициент"""
    petrophysic_selected_dict = {}
    for i, key_p in enumerate(Inform_for_method.Types_task):
        if key_p in petrophysic_selected:
            petr_value = PETROPHYSICS[key_p]
        else:
            petr_value = '-'
        petrophysic_selected_dict[f"petrophysic_selected_{i + 1}"] = petr_value
    return petrophysic_selected_dict


def for_method_data(data, key_m):
    """для каждого метода записывает данные"""
    method = []
    for method_info in data[key_m]:
        for second_table_info in data["second_table"]:
            if method_info["id"] == second_table_info["inform_for_method_id"]:
                method_info.update(second_table_info)

        method_info["data_type"] = data["data_type"]
        method_info.update(petrophysic_selected_num(method_info["petrophysic_selected"]))
        method_info["petrophysic_selected"] = get_petrophysic_string(method_info["petrophysic_selected"])
        method.append(method_info)
    return method


def methods_data(key_m, data, height_descr_list):
    """формирует информацию по методам и их характеристике"""
    n_data = {}
    if key_m == "inform_for_method":
        n_data.update({"method": for_method_data(data, key_m), "numb_method": len(data[key_m])})

        # new_data.update({"author": creater})
        descr_list, height_descr_list = get_descriptions(data)
        n_data.update({"reason": descr_list})
    return n_data, height_descr_list


def get_type_list(data_type):
    """формирует информацию о типе отчета"""
    type_list = "Оперативная оценка"
    if data_type == "Из памяти прибора":
        type_list = "Финальный отчет"
    return type_list


def get_name_qual(field_name, cluster_name, num_well):
    """формирует название отчета"""
    return f"Общая оценка качества каротажа месторождения {field_name}, " \
           f"куста {cluster_name}, скважины {num_well} "


def new_data_height(new_data, height_descr_list):
    """формирует данные о высоте ячейки"""
    for key in new_data.keys():
        if key == "method":
            for method in new_data[key]:
                for key_method in method.keys():
                    method[key_method] = {"value": method[key_method], "height": None}
        else:
            new_data[key] = {"value": new_data[key], "height": None}
            if key == "reason":
                new_data[key]["height"] = height_descr_list
    return


def get_style_key(style_key, data, special_keys):
    """записывает информацию о стиле ячейки"""
    return {"border": None,
            "alignment": Alignment(horizontal=get_horizontal(style_key),
                                   vertical='center', wrapText=True),
            "font": Font(name='Segoe UI', size=get_size(style_key),
                         color=get_color(style_key, data[style_key]["value"],
                                         special_keys),
                         bold=get_bold(style_key), italic=get_italic(style_key)),
            "number_format": get_number_format(style_key,
                                               data[style_key]["value"])}


def get_style_data(new_data, special_keys):
    """записывает информацию о стиле ячеек"""
    for key in new_data.keys():
        if key == "method":
            for method in new_data[key]:
                for key_method in method.keys():
                    method[key_method].update(get_style_key(key_method, method, special_keys))
                    # method[key_method].update({"border": None,
                    #                            "alignment": Alignment(horizontal=get_horizontal(key_method),
                    #                                                   vertical='center', wrapText=True),
                    #                            "font": Font(name='Segoe UI', size=get_size(key_method),
                    #                                         color=get_color(key_method, method[key_method]["value"],
                    #                                                         special_keys),
                    #                                         bold=get_bold(key_method), italic=get_italic(key_method)),
                    #                            "number_format": get_number_format(key_method,
                    #                                                               method[key_method]["value"])})
        else:
            new_data[key].update(get_style_key(key, new_data, special_keys))
            # new_data[key].update({"border": None,
            #                       "alignment": Alignment(horizontal=get_horizontal(key), vertical='center',
            #                                              wrapText=True),
            #                       "font": Font(name='Segoe UI', size=get_size(key),
            #                                    color=get_color(key, new_data[key]["value"], special_keys),
            #                                    bold=get_bold(key), italic=get_italic(key)),
            #                       "number_format": get_number_format(key, new_data[key]["value"])})
    return


def get_data_dict(data, creater):
    """Формирует словарь для дальнейшего преобразования и записи"""
    new_data = get_without_spec(data)
    height_descr_list = 60
    special_keys = []

    for key in data.keys():
        special_keys = special_keys + special_data(key, data, new_data)

        method_dict, height_descr_list = methods_data(key, data, height_descr_list)
        new_data.update(method_dict)

        new_data.update({"data_type_header": get_type_list(data["data_type"]),
                         "name_quality_list": get_name_qual(data['field_name'],
                                                            data['cluster_name'], data['num_well'])})
    new_data_height(new_data, height_descr_list)

    get_style_data(new_data, special_keys)
    del new_data["data_type"]

    return new_data


def status_false(value):
    """Возвращает статус"""
    return "Нет"


def value_not_pers(value):
    """перевод из процентов в числовой формат"""
    if value and type(value) is not str:
        return value / 100


def value_count(data, func):
    """запись информации в данные по ключам"""
    for key in data:
        if "count" in key:
            data[key] = func(data[key])
    return data


def value_count_status(data):
    """вызывает в зависимости от статуса функцию и отправляет ее в функцию записи информации в данные"""
    if "status" in data.keys():
        if data["status"] == "-":
            data = value_count(data, status_false)
            return data
    data = value_count(data, value_not_pers)
    return data


def get_color(key, value, special_keys):
    """определяет цвет для ячейки"""
    if key in special_keys:
        color = get_general_color(value)
        return color
    color = get_color_string(value)
    if value is None or value == "-":
        return "000000"
    if not color:
        if key == "koef_shod":
            for color in COLOR:
                if COLOR[color][0] * 100 <= value <= COLOR[color][1] * 100:
                    return color
        if key == "koef_fail":
            for color in COLOR_FAIL:
                if COLOR_FAIL[color][0] * 100 <= value <= COLOR_FAIL[color][1] * 100:
                    return color
        return "000000"
    else:
        return color


def get_color_string(value):
    """определяет цвет для строкового значения"""
    for color in COLOR_STRING:
        if value in COLOR_STRING[color]:
            return color
    return


def get_general_color(value):
    """определяет цвет для данных full_inform/digital"""
    for color in COLOR_STRING_GENERAL:
        if type(value) == str:
            if value in COLOR_STRING_GENERAL[color]:
                return color
    return "000000"


def get_horizontal(key):
    """определяет horizontal для ячейки"""
    if key in ["note", "reason", "data_status"]:
        return "left"
    return "center"


def get_size(key):
    """определяем значения для размера текста (по ключами значений)"""
    if key in ["value", "author"]:
        return 12
    elif key == "data_type_header":
        return 20
    elif key in ["name_quality_list", "note", "reason"]:
        return 11
    return 9


def get_number_format(key, value):
    """определяем значения для числового формата (по ключами значений: проценты или числовой)"""
    if key in ["full_inf_count", "digital_count"] and value != "Нет":
        return "0.00%"
    elif key in ["section_interval_start", "section_interval_end", "diametr", "petrophysic_selected_1",
                 "petrophysic_selected_2", "petrophysic_selected_3", "petrophysic_selected_4", "petrophysic_selected_5",
                 "interval_shod_start", "interval_shod_end", "value"]:
        return "0.00"
    elif key in ["koef_shod", "koef_fail", "method_value"]:
        return "0"
    elif key in ["start_date", "end_date", "calibr_date"]:
        return "DD.MM.YYYY"
    else:
        return "General"


def get_italic(key):
    """Определяем значения для italic"""
    if key in ["full_inf_count", "digital_count",
               "linkage", "emissions", "noise", "control", "distribute_support", "distribute_palet", "dash",
               "corresponse", "correlation", "device_tech_condition", "method_value", "notes",
               "petrophysic_selected_1", "petrophysic_selected_2", "petrophysic_selected_3",
               "petrophysic_selected_4", "petrophysic_selected_5"]:
        return True
    return False


def get_bold(key):
    """определяем значения для bold"""
    if key in ["value", "name_quality_list", "data_status"]:
        return True
    return False


def work_wb(wb, value_data):
    """Работа с листами"""
    for sheet in wb:
        work_sheet(sheet, value_data)


def create_excel_dict(value_data_copy):
    """Формирует словарь с ключами из шаблона excel"""
    value_data_excel = {}
    for key_value in value_data_copy.keys():
        value_data_excel[f"{{{{{key_value}}}}}"] = value_data_copy[key_value]
        if key_value == "method":
            value_method_excel = []
            value_method = copy.deepcopy(value_data_copy[key_value])
            for method in value_method:
                new_method = {}
                for key_method in method.keys():
                    new_method[f"{{{{{key_method}}}}}"] = method[key_method]
                value_method_excel.append(new_method)
            value_data_excel[f"{{{{{key_value}}}}}"] = value_method_excel
    return value_data_excel


def get_merged_ranges(sheet, cell, merged_ranges):
    """получение информации об ячейках, которые необходимо сначала разъединить, а далее слить заново"""
    if type(cell) == MergedCell:
        if cell.column > 1:

            if type(sheet.cell(row=cell.row, column=cell.column - 1)) == Cell:
                merged_ranges.append({"cell_start": sheet.cell(row=cell.row, column=cell.column - 1),
                                      "cell_end": None})
            if len(merged_ranges) > 0:
                if merged_ranges[-1]["cell_start"] is not None and \
                        merged_ranges[-1]["cell_end"] is None and \
                        type(sheet.cell(row=cell.row, column=cell.column + 1)) == Cell:
                    merged_ranges[-1]["cell_end"] = cell
    return merged_ranges


def not_accord_method(value_data_excel, cell, cell_height):
    """заполнение в ячейку данных не по методам"""
    if cell.value in value_data_excel.keys():

        cell.font = value_data_excel[cell.value]["font"]
        cell.number_format = value_data_excel[cell.value]["number_format"]
        cell.alignment = value_data_excel[cell.value]["alignment"]
        if value_data_excel[cell.value]["height"] is not None:
            cell_height.append({"cell": cell, "height": value_data_excel[cell.value]["height"]})

        cell.value = value_data_excel[cell.value]["value"]
    return cell_height


def set_not_method(value_data_excel, sheet):
    """Заполнение в лист данных не по методам"""
    method_cell = None
    merged_ranges = []
    cell_height = []
    for row in sheet.rows:
        for cell in row:
            merged_ranges = get_merged_ranges(sheet, cell, merged_ranges)

            if cell.value == "{{method}}":
                method_cell = cell
            else:
                cell_height = not_accord_method(value_data_excel, cell, cell_height)
    return method_cell, merged_ranges, cell_height


def unmerge_ranges(sheet, merged_ranges):
    """Разъединение всех ячеек перед созданием новых под методы"""
    for merged_range in merged_ranges:
        sheet.unmerge_cells(f"{get_column_letter(merged_range['cell_start'].column)}"
                            f"{merged_range['cell_start'].row}:"
                            f"{get_column_letter(merged_range['cell_end'].column)}"
                            f"{merged_range['cell_end'].row}")
        merged_range['cell_end'] = sheet.cell(row=merged_range['cell_end'].row, column=merged_range['cell_end'].column)
    return


def set_method(sheet, method_data, method_cell):
    """заполнение данных по методам"""
    numb_method = len(method_data)
    if method_cell is not None:
        for i, method in enumerate(method_data):
            if i < (len(method_data) - 1):
                sheet.insert_rows(method_cell.row + i + 1)
            row_numb = method_cell.row + i
            for row in sheet.iter_rows(min_row=row_numb, max_row=row_numb):
                for cell in row:
                    if i < (len(method_data) - 1):
                        copy_method_cell(sheet, cell, row_numb)
                    if cell.value in method.keys():
                        set_method_cell(cell, method)
    return numb_method


def copy_method_cell(sheet, cell, row_numb):
    """копирование информации в последующую ячейку"""
    sheet.row_dimensions[row_numb + 1].height = sheet.row_dimensions[cell.row].height
    new_cell = sheet[f"{get_column_letter(cell.column)}{row_numb + 1}"]
    set_new_cell(new_cell, cell)
    return


def set_new_cell(new_cell, cell):
    """копирование стиля/значения в новую ячейку (созданную ниже для дальнейших методов)"""
    new_cell.value = cell.value
    new_cell.font = cell.font._StyleProxy__target
    new_cell.alignment = cell.alignment._StyleProxy__target
    new_cell.fill = cell.fill._StyleProxy__target
    new_cell.border = cell.border._StyleProxy__target
    new_cell.number_format = cell.number_format
    return


def set_method_cell(cell, method):
    """установка данных по методам"""
    cell.font = method[cell.value]["font"]
    cell.number_format = method[cell.value]["number_format"]
    cell.alignment = method[cell.value]["alignment"]
    cell.value = method[cell.value]["value"]
    return


def empty_data(sheet):
    """заполнение пустых данных при их отсутствии"""
    for row in sheet.rows:
        for cell in row:
            value = cell.value
            if type(value) is str:
                if value == "{{data_status}}":
                    cell.value = "Полнота предоставления цифровых данных"
                if re.search(r'{{\w+}}', value):
                    cell.value = None
    return


def work_sheet(sheet, value_data):
    """Работа с листом"""
    value_data_copy = copy.deepcopy(value_data)
    value_data_excel = create_excel_dict(value_data_copy)
    method_cell, merged_ranges, cell_height = set_not_method(value_data_excel, sheet)

    # unmerge_ranges(sheet, merged_ranges)

    numb_method = 0
    try:
        numb_method = set_method(sheet, value_data_excel["{{method}}"], method_cell)
    except KeyError:
        pass
    empty_data(sheet)

    merge_ranges(numb_method, method_cell, merged_ranges, sheet)

    for cell in cell_height:
        sheet.row_dimensions[cell["cell"].row].height = cell["height"]


def merge_ranges(numb_method, method_cell, merged_ranges, sheet):
    """слияние всхе необходимых диапазонов ячеек"""
    for merged_range in merged_ranges:
        if method_cell is not None:
            if merged_range['cell_start'].row == method_cell.row:
                for i in range(1, numb_method + 1):
                    merge_cells_ranges(merged_range, sheet, i)
        merge_cells_ranges(merged_range, sheet)
    return


def merge_cells_ranges(merged_range, sheet, numb=0):
    """слияние конкретного диапазона ячеек"""
    sheet.merge_cells(f"{get_column_letter(merged_range['cell_start'].column)}"
                      f"{merged_range['cell_start'].row + numb}:"
                      f"{get_column_letter(merged_range['cell_end'].column)}"
                      f"{merged_range['cell_end'].row + numb}")
    return


def save_xlsx(data, pk, creater):
    """Сохранение отчета в xlsx"""
    wb = get_template()
    if wb == "There is no template":
        return 'Error: Нет доступных шаблонов отчёта'
    # sheets = get_sheets(wb)

    value_data = get_data_dict(data, creater)
    work_wb(wb, value_data)

    create_name_file = f'{data["customer_name"]}_{data["field_name"]}_{data["num_pad"]}_{data["num_well"]}_{pk}'
    name_file = '_'.join(create_name_file.split())
    wb.save(os.getcwd() + f'/check_list/src/files_root/report_file/{name_file}.xlsx')

    wb.close()
    return name_file


def get_template():
    """находим template"""
    template = Sample.objects.last()
    if template:
        file = template.sample_file
        return load_workbook(file)
    else:
        return "There is no template"


def save_pdf(data, pk, creater):
    """Сохранение отчета в pdf"""
    name_file = save_xlsx(data, pk, creater)

    # df = pd.read_excel(f'files_root\\checklist\\{name_file}.xlsx')
    # df.to_html(f'files_root\\checklist\\{name_file}.html')
    # config = pdfkit.configuration(wkhtmltopdf='files_root\\wkhtmltopdf\\bin\\wkhtmltopdf.exe')
    # pdfkit.from_file(f'files_root\\checklist\\{name_file}.html', f'files_root\\checklist\\{name_file}.pdf',
    #                  configuration=config)
    if platform.system() == 'Windows':
        book = xw.Book(os.getcwd() + f'\\check_list\\src\\files_root\\report_file\\{name_file}.xlsx')
        book.to_pdf(os.getcwd() + f'\\check_list\\src\\files_root\\report_file\\{name_file}.pdf')
        app = book.app
        app.visible = False
        app.kill()
    else:
        os.system(f'libreoffice --headless --convert-to pdf check_list/src/files_root/report_file/{name_file}.xlsx '
                  f'--outdir  check_list/src/files_root/report_file/')

    return name_file
