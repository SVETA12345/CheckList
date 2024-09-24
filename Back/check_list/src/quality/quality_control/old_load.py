from openpyxl import load_workbook
from enum import Enum
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from quality.inform_for_method.models import Inform_for_method

import xlwings as xw


# список названий листов
class SheetNames(Enum):
    SHEET_1 = "Лист1. Общая информация"
    SHEET_2 = "Лист2. Оценка качества ГИС"
    SHEET_3 = "Лист3. Оценка петрозадачи"
    SHEET_4 = "Лист4. Графика"


# петрофизические задачи с распределенными коэффициентами
PETROPHYSICS_VALUE = [0.35, 0.25, 0.20, 0.10, 0.10]
PETROPHYSICS = {}
DEFAULT = 15
AUTHOR = ""
for i, key in enumerate(Inform_for_method.Types_task):
    PETROPHYSICS.update({key: PETROPHYSICS_VALUE[i]})
WIDTH = [27.00, 16.00, 16.00, 16.00, 16.00, 16.00, 23.00, 23.00, 23.00, 23.00, 23.00, 23.00, 11.00, 40.00]

keys_full_inform = ["titul_list", "well_construction", "wellbore_sizes", "chrono_data", "sol_data", "dash_comp",
                    "summary_data",
                    "inklino_data", "main_record", "parametr", "control_record", "lqc", "calibration"]

keys_las_file = ["cap", "parametres", "mnemodescription", "tabledata"]

keys_witsml = ["fullness_data", "curvenames", "mnemodescription"]

# цвета и соответствующие им значения: string
COLOR_STRING = {
    "00B050": ["Имеется", "Увязан", "Отсутствуют", "Отсутствует", "Соответствует основному замеру",
               "Соответствует данным опорных скважин", "Лежат в области ожидаемых значений", "Соответствуют",
               "Хорошее", "Коррелируют"],
    "0000FF": ["Частично неувязан", "Единичные", "Незначительная", "Частично коррелируют",
               "Отменена по согласованию сторон", "Не соответствует основной записи",
               "Не соответствует данным опорных скважин (занижено)", "Завышены",
               "Реперные горизонты не вскрыты", "Свойства реперного горизонта отсутствуют",
               "Не соответствует данным опорных скважин (завышено)", "Занижены", "Удовлетворительное"],
    "FF0000": ["Не имеется", "Неувязан", "Регулярные", "Высокая",
               "Не произведена", "Не лежат в области ожидаемых значений", "Не соответствуют",
               "Свойства реперного горизонта расходятся", "Неудовлетворительное", "Не коррелируют"]}

COLOR_STRING_GENERAL = {
    "00B050": "Полная", "0000FF": "Частичная", "FF0000": "Отсутствует"
}

# цвета и соответствующие им значения: числа
COLOR = {"00B050": [0.85, 1.00], "0000FF": [0.65, 0.84], "FF0000": [0.00, 0.64]}
COLOR_FAIL = {"00B050": [0.00, 0.10], "0000FF": [0.10, 0.20], "FF0000": [0.20, 1.00]}
# значения для первого листа для расчетных значений по каждой отдельной таблице
COUNT = ["M15", "M22", "M26"]


# работа со стилями
class Style:
    # значение ячейки
    def __init__(self, value):
        self.value = value

    # определяем цвет для строкового значения
    def get_color_string(self):
        for color in COLOR_STRING:
            if self.value in COLOR_STRING[color]:
                return color
        return

    def get_general_color(self):
        for color in COLOR_STRING_GENERAL:
            if self.value in COLOR_STRING_GENERAL[color]:
                return color
        return

    # определяем цвет для значения (строковое + числовое)
    def get_color(self, key):
        keys = keys_las_file + keys_full_inform + keys_witsml
        if key in keys:
            color = self.get_general_color()
            return color
        color = self.get_color_string()
        if self.value is None or self.value == "-":
            return "000000"
        if not color:
            if key == "koef_shod":
                for color in COLOR:
                    if COLOR[color][0]*100 <= self.value <= COLOR[color][1]*100:
                        return color
            if key == "koef_fail":
                for color in COLOR_FAIL:
                    if COLOR_FAIL[color][0]*100 <= self.value <= COLOR_FAIL[color][1]*100:
                        return color
            return "000000"
        else:
            return color

    # определяем значения для bold (по ключами значений)
    def get_bold(self, key):
        if key in ["full_inf_count", "witsml_count", "las_file_count", "value", "name_quality_list"]:
            return True
        return False

    # определяем значения для размера текста (по ключами значений)
    def get_size(self, key):
        if key in ["full_inf_count", "witsml_count", "las_file_count"]:
            return 10
        elif key == "value":
            return 12
        elif key == "type":
            return 20
        elif key in ["name_quality_list", "data_of_created", "description"]:
            return 11
        return 9

    # определяем значения для текста (по ключами значений + все параметры)
    def get_font(self, italic, key):
        if key in ["full_inf_count", "witsml_count", "las_file_count", "method"]:
            italic = False
        return Font(name='Segoe UI', size=self.get_size(key), color=self.get_color(key), bold=self.get_bold(key),
                    italic=italic)

    # определяем значения для alignment (по ключами значений)
    def get_alignment(self, alignment):
        return Alignment(horizontal=alignment, vertical='center', wrapText=True)

    # определяем значения для числового формата (по ключами значений: проценты или числовой)
    def get_number_format(self, number_format, key):
        if key in ["full_inf_count", "witsml_count", "las_file_count"]:
            return "0.00%"
        else:
            return number_format


# заполняем заданные в key клетки данными/стилем
def data_in_cell(ws, data):
    for key in data:
        ws[key] = data[key]["value"]
        ws[key].alignment = data[key]["alignment"]
        ws[key].font = data[key]["font"]
        ws[key].number_format = data[key]["number_format"]


# формируем элемент таблицы: каждой клетке соответствует ряд параметров
def get_element_dict(cell, value, font, alignment, number_format):
    return {cell: {"value": value, "font": font, "alignment": alignment, "number_format": number_format}}


# формируем ряд элементов таблицы из исходных данных, клеток, keys - ключей передаваемого json
def get_data_keys(data, keys, cells, alignment="center", italic=False, number_format="General"):
    data_keys = {}
    for numb in range(len(cells)):
        try:
            inf_cell = data[keys[numb]]
            style = Style(inf_cell)
            data_keys.update(get_element_dict(cells[numb], inf_cell, style.get_font(italic, keys[numb]),
                                              style.get_alignment(alignment),
                                              style.get_number_format(number_format, keys[numb])))
        except KeyError:
            pass
    return data_keys


# определяем тип отчета и возвращаем в формате таблицы с ключем-ячейкой
def get_type(data_type):

    type_list = "Оперативная оценка"
    if data_type == "Из памяти прибора":
        type_list = "Финальный отчет"
    return get_element_dict("A5", type_list, Style(data_type).get_font(False, "type"),
                            Style(data_type).get_alignment("center"), "General")


# формируем dict данных для общих данных листов (первая таблица с основной информацией)
def get_general_data(data):
    dict_data = get_type(data["data_type"])
    cells = ["A9", "B9", "E9", "F9", "G9", "H9", "I9", "K9", "L9", "M9"]
    keys = ["customer_name", "field_name", "num_pad", "num_well", "well_type", "pie_well", "service_name",
            "section_interval_start", "section_interval_end", "diametr"]
    dict_data.update(get_data_keys(data, keys, cells))

    return dict_data


# работа со слиянием ячеек для общей таблицы на листе
def work_cell_merge_general(ws):
    ws.merge_cells('B9:D9')
    ws.merge_cells('I9:J9')
    ws.merge_cells('M9:N9')


# расчет высоты ячейки по заданным параметрам (transfer - количество строк фразы,
# 5 - просчет для неровного деления слов, transfer + 1 - количество расстояний между фразами, 4 - размер расстояния)
# def get_value_height(size, current_width, value):
#     mul = 0
#     for v in str(value).split('\n'):
#         transfer = math.ceil(len(v) / current_width) + 1
#         mul += transfer * (size + 4) + (transfer + 1) * (4 / transfer)
#     return mul


# определяет для строки наибольший размер с учетом слитых ячеек
# def auto_height(ws, row_number):
#     for row in ws.iter_rows(min_row=row_number, min_col=1, max_col=13, max_row=row_number):
#         default_height = 15
#         multiples_of_font_size = [default_height]
#         merge_multiples_of_font_size = []
#         value = ""
#         size = 0
#         merge_cell = 0
#         count_merge = 0
#         default_height = 15
#         for j, cell in enumerate(row):
#             col_letter = get_column_letter(cell.column)
#             current_width = ws.column_dimensions[col_letter].width
#
#             if type(cell) is not MergedCell and cell.value is not None:
#                 value = cell.value
#                 size = cell.font.size
#                 mul = get_value_height(size, current_width, value)
#                 if mul > 0:
#                     multiples_of_font_size.append(mul)
#                 merge_cell = 0
#                 count_merge = 0
#                 merge_multiples_of_font_size = []
#             else:
#                 merge_cell = merge_cell + current_width
#                 multiples_of_font_size.pop()
#                 count_merge = count_merge + 1
#                 mul = get_value_height(size, merge_cell, value)
#
#                 if mul > 0:
#                     merge_multiples_of_font_size.append(mul)
#                     multiples_of_font_size.append(min(merge_multiples_of_font_size))
#
#         original_height = ws.row_dimensions[row_number].height
#         if original_height is None:
#             original_height = default_height
#         new_height = max(multiples_of_font_size)
#         if original_height < new_height:
#             ws.row_dimensions[row_number].height = new_height


# установка лого
def create_header_img(ws):
    img = Image('files_root\\logo.png')
    ws.add_image(img, 'L1')


# установка ширины
def work_cell_width(ws):
    for column in range(1, 10):
        ws.column_dimensions[get_column_letter(column)].width = WIDTH[column - 1]


# установка высоты для общей части
def work_cell_height_general(ws):
    for row in range(11):
        if row == 8:
            ws.row_dimensions[row].height = 45
        elif row == 9:
            ws.row_dimensions[row].height = 30
        else:
            ws.row_dimensions[row].height = DEFAULT


# заполняем основную таблицу на заданном листе
def create_header(ws, gen_data):
    # work_cell_width(ws)

    # work_cell_merge_general(ws)
    data_in_cell(ws, gen_data)

    # create_header_img(ws)
    # auto_height(ws, 9)


# заполнение основной таблицы на листах книги
def general_work_ws(wb, gen_data):
    for sheet in wb:
        create_header(sheet, gen_data)


# получение строки информации о задачах
def get_petrophysic_string(petrophysic_selected, ):
    petrophysic_string = ""
    amount_petrophysic = len(petrophysic_selected) - 1
    if len(petrophysic_selected) > 0:
        for i in range(amount_petrophysic):
            petrophysic_string = petrophysic_string + f"{petrophysic_selected[i]};\n"
        petrophysic_string = petrophysic_string + f"{petrophysic_selected[amount_petrophysic]}."
    return petrophysic_string


# формирование dict для информации для одного метода для первого листа (основная инормация)
def data_method(numb_row, method):
    cells = [f"A{numb_row}", f"B{numb_row}", f"C{numb_row}",
             f"G{numb_row}", f"H{numb_row}", f"J{numb_row}", f"K{numb_row}", f"L{numb_row}", f"N{numb_row}"]
    keys = ["method", "tool_type", "tool_num", "interval_shod_start",
            "interval_shod_end", "koef_shod", "reason_rashod", "koef_fail", "petrophysic_string"]
    method["petrophysic_string"] = get_petrophysic_string(method["petrophysic_selected"])

    return get_data_keys(method, keys, cells)


# формирование dict для информации по всем методам для первого листа (основная инормация по каждому методу)
def data_methods(first_row, inform_for_method, data_type):
    methods_inform = {}
    numb_row = first_row
    for method in inform_for_method:
        methods_inform.update(data_method(numb_row, method))
        methods_inform.update(get_element_dict(f"I{numb_row}", data_type, Style(data_type).get_font(False, "data_type"),
                                               Style(data_type).get_alignment("center"), "General"))
        methods_inform.update(
            get_element_dict(f"D{numb_row}", method["calibr_date"], Style(method["calibr_date"]).get_font(False, "calibr_date"),
                                               Style(method["calibr_date"]).get_alignment("center"), "DD.MM.YYYY"))
        methods_inform.update(
            get_element_dict(f"E{numb_row}", method["start_date"],
                             Style(method["start_date"]).get_font(False, "start_date"),
                             Style(method["start_date"]).get_alignment("center"), "DD.MM.YYYY"))
        methods_inform.update(
            get_element_dict(f"F{numb_row}", method["end_date"], Style(method["end_date"]).get_font(False, "end_date"),
                             Style(method["end_date"]).get_alignment("center"), "DD.MM.YYYY"))
        methods_inform.update(
            get_element_dict(f"M{numb_row}", method["method_value"],
                             Style(method["method_value"]).get_font(True, "method_value"),
                             Style(method["method_value"]).get_alignment("center"), "0"))
        # f"D{numb_row}", f"E{numb_row}", f"F{numb_row}"
        # "calibr_date", "start_date", "end_date",
        numb_row = numb_row + 1
    return methods_inform


# получаем информацию по методам и количество методов
def get_main_data_sh1(data, first_row):
    inform_for_method = data["inform_for_method"]
    amount_method = len(inform_for_method)

    return data_methods(first_row, inform_for_method, data["data_type"]), amount_method


# заполняем клетки, сформированные по количеству методов, стилем
def style_row(ws, amount_method):
    for i in range(1, amount_method + 1):
        row = i + 11
        for row in ws.iter_rows(min_row=row, min_col=1, max_col=14, max_row=row):
            for cell in row:
                cell.border = Border(left=Side(border_style="medium", color='FFFFFF'),
                                     right=Side(border_style="medium", color='FFFFFF'),
                                     top=Side(border_style="medium", color='FFFFFF'),
                                     bottom=Side(border_style="medium", color='FFFFFF'))
                cell.fill = PatternFill("solid", fgColor="faf0c8")


# установка высоты в меняющихся таблицах
def main_cell_height(ws, amount_method):
    for row in range(11, 12 + amount_method):
        if row == 11:
            ws.row_dimensions[row].height = 45
        else:
            ws.row_dimensions[row].height = 75


# основная работа с первым листом: header + таблица методов
def main_work_sh1(ws, data):
    methods_inform, amount_method = get_main_data_sh1(data, 12)
    ws.move_range("A13:N43", rows=(amount_method - 1), cols=0, translate=True)
    style_row(ws, amount_method)

    data_in_cell(ws, methods_inform)
    # set_auto_height(ws, amount_method)
    main_cell_height(ws, amount_method)

    return amount_method


# слияния клеток для таблицы с полнотой предоставленных данных
def first_table_sh1(ws, begin_row):
    ws.merge_cells(f'A{begin_row}:M{begin_row}')

    end_row = begin_row + 5
    for row in range(begin_row, end_row):
        ws.merge_cells(f'A{row}:D{row}')
        ws.merge_cells(f'F{row}:H{row}')
        ws.merge_cells(f'J{row}:M{row}')

    ws.merge_cells(f'A{end_row}:D{end_row}')
    ws.merge_cells(f'F{end_row}:I{end_row}')
    ws.merge_cells(f'J{end_row}:N{end_row}')

    return end_row + 2


# слияния клеток для таблицы с оформление las-файла
def second_table_sh1(ws, begin_second_table):
    ws.merge_cells(f'A{begin_second_table}:M{begin_second_table}')

    end_row = begin_second_table + 3
    for row in range(begin_second_table, end_row):
        ws.merge_cells(f'A{row}:G{row}')
        ws.merge_cells(f'I{row}:M{row}')

    return end_row + 1


# слияния клеток для таблицы с корректностью witsml
def third_table_sh1(ws, begin_third_table):
    ws.merge_cells(f'A{begin_third_table}:M{begin_third_table}')

    row_1 = begin_third_table + 1
    ws.merge_cells(f'A{row_1}:F{row_1}')
    ws.merge_cells(f'H{row_1}:M{row_1}')

    row_2 = row_1 + 1
    ws.merge_cells(f'A{row_2}:F{row_2}')
    ws.merge_cells(f'H{row_2}:N{row_2}')

    return row_2 + 2


# слияния клеток для таблицы с полнотой данных реального времени
def fourth_table_sh1(ws, begin_fourth_table):
    ws.merge_cells(f'A{begin_fourth_table}:M{begin_fourth_table}')
    row = begin_fourth_table + 1
    ws.merge_cells(f'A{row}:F{row}')
    ws.merge_cells(f'H{row}:N{row}')
    return row


# слияния клеток для четырех таблиц первого листа после методов
def work_cell_merge_sh1(ws, begin_first_table):
    begin_second_table = first_table_sh1(ws, begin_first_table)
    begin_third_table = second_table_sh1(ws, begin_second_table)
    begin_fourth_table = third_table_sh1(ws, begin_third_table)
    end_fourth_table = fourth_table_sh1(ws, begin_fourth_table)
    end_merge_sh1(ws, end_fourth_table + 6)


def end_merge_sh1(ws, end_merge_row):
    ws.merge_cells(f'A{end_merge_row}:N{end_merge_row}')
    ws.merge_cells(f'A{end_merge_row+1}:N{end_merge_row+1}')
    ws.merge_cells(f'A{end_merge_row+2}:N{end_merge_row+2}')


# формирование названий клеток для информации полноты предоставляемых данных
def get_cells_full_inform(begin_row, letter):
    cells = []
    for number in range(4):
        cells = cells + [f"{letter}{begin_row + number}"]
    return cells


# формирование dict для информации полноты предоставляемых данных
def full_inform(full_inform_data, begin_row):
    cells = get_cells_full_inform(begin_row, "E") + [f"E{begin_row + 4}"] + get_cells_full_inform(begin_row, "I") + \
            get_cells_full_inform(begin_row, "N") + [f"N{begin_row - 1}"]

    value_count(full_inform_data, value_not_pers)

    return get_data_keys(full_inform_data, keys_full_inform + ["full_inf_count"], cells, "left", True)


# формирование dict для информации оформления las-файла
def las_file(las_file_data, begin_row):
    cells = [f"H{begin_row}", f"H{begin_row + 1}", f"N{begin_row}", f"N{begin_row + 1}", f"N{begin_row - 1}"]

    las_file_data = not_used_data(las_file_data, keys_las_file + ["las_file_count"])
    las_file_data = value_count_status(las_file_data)

    return get_data_keys(las_file_data, keys_las_file + ["las_file_count"], cells, "left", True)


# в ячейки с отсутсвтующими данными записывается -
def not_used_data(data, keys):
    if data["status"] is False:
        for key in keys:
            data[key] = "-"
    return data


# формирование dict для информации корректности witsml
def witsml(witsml_data, begin_row):
    cells = [f"G{begin_row}", f"G{begin_row + 1}", f"N{begin_row}", f"N{begin_row - 1}"]

    witsml_data = not_used_data(witsml_data, keys_witsml + ["witsml_count"])
    witsml_data = value_count_status(witsml_data)

    return get_data_keys(witsml_data, keys_witsml + ["witsml_count"], cells, "left", True)


# формирование dict для информации полноты данных реального времени
def density(density_data, begin_row):
    cells = [f"G{begin_row}"]
    keys = ["density"]
    return get_data_keys(density_data, keys, cells, "left", True)


def status_false(value):
    return "Нет"


# перевод из процентов в числовой формат
def value_not_pers(value):
    if value and type(value) is not str:
        return value / 100


# запись информации в данные по ключам
def value_count(data, func):
    for key in data:
        if "count" in key:
            data[key] = func(data[key])
    return data


# вызывает в зависимости от статуса функцию и отправляет ее в функцию записи информации в данные
def value_count_status(data):
    if data["status"] is False:
        data = value_count(data, status_false)
    else:
        data = value_count(data, value_not_pers)
    return data


# получение таблицы данных для четырех нижних таблиц первого листа
def get_particular_inf(data, begin_row):
    particular_inf = {}

    particular_inf.update(full_inform(data["full_inform"], begin_row))
    particular_inf.update(las_file(data["las_file"], begin_row + 7))
    particular_inf.update(witsml(data["witsml"], begin_row + 11))
    particular_inf.update(density(data["density"], begin_row + 15))
    return particular_inf


# установка DEFAULT высоты ячейкам
def work_particular_height(ws, begin_row):
    for row in range(begin_row - 1, 23 + begin_row):
        ws.row_dimensions[row].height = DEFAULT


# работа с четырьмя нижними таблицами первого листа
def work_particular_sh1(ws, begin_row, data, creater):
    methods_inform = get_particular_inf(data, begin_row)
    methods_inform.update(get_name_quality_list(begin_row + 22, data))
    methods_inform.update(get_element_dict(f"N{begin_row + 27}", data["value"],
                                           Style(data["value"]).get_font(False, "value"),
                                           Style(data["value"]).get_alignment("center"), "General"))
    methods_inform.update(get_element_dict(f"N{begin_row + 28}", creater,
                                           Style(AUTHOR).get_font(True, "data_of_created"),
                                           Style(AUTHOR).get_alignment("center"), "General"))

    data_in_cell(ws, methods_inform)


# работа с первым листом
def work_sheet_1(ws, data, creater):
    amount_method = main_work_sh1(ws, data)

    begin_row = amount_method + 13

    work_particular_height(ws, begin_row)
    work_particular_sh1(ws, begin_row + 1, data, creater)

    work_cell_merge_sh1(ws, begin_row)


# формирование dict для информации оценки качества метода
def second_table_data(numb_row, second_table):
    cells = [f"A{numb_row}", f"B{numb_row}", f"C{numb_row}", f"D{numb_row}", f"E{numb_row}", f"F{numb_row}",
             f"G{numb_row}", f"H{numb_row}", f"I{numb_row}", f"J{numb_row}", f"K{numb_row}", f"L{numb_row}",
             f"M{numb_row}", f"N{numb_row}"]
    keys = ["method", "act", "linkage", "emissions", "noise", "control", "distribute_support",
            "distribute_palet", "dash", "corresponse", "correlation", "device_tech_condition", "method_value",
            "notes"]
    return get_data_keys(second_table, keys, cells, italic=True)


# формирование dict для информации оценки качества методов
def table_data_sh2(first_row, second_tables):
    second_tables_inf = {}
    numb_row = first_row
    for second_table in second_tables:
        second_tables_inf.update(second_table_data(numb_row, second_table))
        numb_row = numb_row + 1
    return second_tables_inf


# получение информации качества методов + их количество
def get_main_data_sh2(data, first_row):
    second_tables = data["second_table"]
    amount_method = len(second_tables)

    return table_data_sh2(first_row, second_tables), amount_method


# основная работа второго листа - работа с качеством методов
def main_work_sh2(ws, data):
    methods_inform, amount_method = get_main_data_sh2(data, 12)
    ws.move_range("A13:N24", rows=(amount_method - 1), cols=0, translate=True)
    style_row(ws, amount_method)

    data_in_cell(ws, methods_inform)

    main_cell_height(ws, amount_method)

    # set_auto_height(ws, amount_method)

    return amount_method


# формирование dict для даты создания отчета (на второй лист)
# def get_data_created(data_end_row, data_of_created):
#     return get_element_dict(f"M{data_end_row + 3}", data_of_created,
#                             Style(data_of_created).get_font(True, "data_of_created"),
#                             Style(data_of_created).get_alignment("left"), "General")


# формирование dict для значения качества
def get_value(begin_row, value):
    return get_element_dict(f"M{begin_row}", value, Style(value).get_font(False, "value"),
                            Style(value).get_alignment("center"), "General")


# формирование dict для описания оценки качества
def get_name_quality_list(data_end_row, data):
    name_quality_list = f"Общая оценка качества каротажа месторождения {data['field_name']}, " \
                        f"куста {data['cluster_name']}, скважины {data['num_well']} "
    return get_element_dict(f"A{data_end_row}", name_quality_list,
                            Style(name_quality_list).get_font(False, "name_quality_list"),
                            Style(name_quality_list).get_alignment("center"), "General")


# работа по слиянию клеток нижней части 2 и 3 листов (информация + причины применения понижающего коэфф)
def merge_end(ws, data_end_row):
    for row in range(data_end_row, data_end_row + 4):
        ws.merge_cells(f'A{row}:N{row}')


# находим соответствие между значением параметра и списком критических значений для понижающего коэфф,
# возвращаем местоположение в list
def get_value_descr_individ(value_parametr):
    values = ["Единичные", "Регулярные", "Незначительная", "Высокая", "Не соответствует", "Частично коррелируют",
              "Не коррелируют", "Низкая плотность данных реального времени",
              "Проблема с передачей данных реального времени", "Перерасчет данных из памяти прибора",
              "Не имеется", "Частично не увязан", "Не увязан",
              "Не соответствует основной записи", "Не произведена",
              "Не соответствует данным опорных скважин (занижено)",
              "Не соответствует данным опорных скважин (завышено)",
              "Не лежат в области ожидаемых значений", "Занижены", "Завышены",
              "Реперные горизонты не вскрыты", "Свойства реперного горизонта отсутствуют",
              "Свойства реперного горизонта расходятся", "Удовлетворительное",
              "Неудовлетворительное"]
    for i in range(len(values)):
        if value_parametr == values[i]:
            return i


# находим по задаваемому местоположению причину применения понижающего коэфф
def get_result_description(name, number):
    results = [f"На данных ГИС {name} при бурении/из памяти прибора наблюдаются единичные выбросы",
               f"На данных ГИС {name} при бурении/из памяти прибора наблюдаются регулярные выбросы",
               f"На данных ГИС {name} наблюдается незначительная зашумленность",
               f"На данных ГИС {name} наблюдается высокая зашумленность",
               f"Показания метода {name} не соответствуют вскрытому разрезу",
               f"{name} частично коррелирует с другими методами ГИС",
               f"{name} не коррелирует с другими методами ГИС",
               "Отмечается низкая плотность данных реального времени",
               "Во время бурения наблюдалась проблема с передачей данных реального времени",
               "Данные из памяти прибора были перерассчитаны", "Отсутствует Акт промера бурового инструмента",
               "Кривые ГИС частично не увязаны по глубине", "Кривые ГИС не увязаны по глубине",
               "Данные контрольной записи не соответствуют основной записи ГИС",
               "Контрольная запись данных ГИС не произведена",
               "Распределение данных ГИС при бурении на гистограммах не соответствует данным опорных скважин (занижено)",
               "Распределение данных ГИС при бурении на гистограммах не соответствует данным опорных скважин (завышено)",
               "Распределение данных ГИС при бурении на кросс-плотах относительно палеток "
               "Не лежит в области ожидаемых значений",
               "Значения абсолютных петрофизических значений в реперных горизонтах занижены",
               "Значения абсолютных петрофизических значений в реперных горизонтах завышены",
               "Реперные горизонты не вскрыты, сравнение абсолютных петрофизических значений "
               "в реперных горизонтах невозможно",
               "Сравнение абсолютных петрофизических значений в реперных горизонтах невозможно, "
               "т.к. свойства реперного горизонта отсуствуют",
               "Сравнение абсолютных петрофизических значений в реперных горизонтах невозможно, "
               "т.к. свойства реперного горизонта расходятся",
               "По показаниям LQC планшетов техническое состояние прибора Удовлетворительное",
               "По показаниям LQC планшетов техническое состояние прибора Неудовлетворительное"]
    return results[number]


# формируем список причин по всем методам
def get_description_list(data):
    second_tables = data["second_table"]
    # part_description = []
    part_description = {}
    comm_names_method = {}
    for method in second_tables:
        for key in method:
            number_result = get_value_descr_individ(method[key])
            if number_result:
                try:
                    name_method = comm_names_method[number_result] + [f'"{method["method"]}"']
                except:
                    name_method = [f'"{method["method"]}"']
                comm_names_method[number_result] = name_method
                part_description[method[key]] = get_result_description(", ".join(set(name_method)), number_result)
    descr_values = list(part_description.values())
    return descr_values, len(descr_values)


# формируем dict с описанием всех причин
def get_descriptions(ws, data, row):
    description_list, amount_row = get_description_list(data)

    ws.row_dimensions[row].height = amount_row * 20

    # description_set = list(set(description_list))
    description = ""
    if len(description_list) != 0:
        len_descr = len(description_list)
        for i in range(len_descr - 1):
            description = description + f"{i + 1}. {description_list[i]};\n"
        description = description + f"{len_descr}. {description_list[len_descr - 1]}."
    return get_element_dict(f"A{row}", description, Style(description).get_font(False, "description"),
                            Style(description).get_alignment("left"), "General")


# работа с информацией внизу 2 и 3 листов
def work_quality_control(ws, begin_row, data):
    data_end = get_value(begin_row + 1, data["value"])

    data_end_row = begin_row + 5

    data_end.update(get_name_quality_list(data_end_row, data))
    data_end.update(get_descriptions(ws, data, data_end_row + 2))
    return data_end, data_end_row


# работа с информацией внизу 2 листа (quality_control + дата создания)
def end_work_sh2(ws, begin_row, data):
    data_end, data_end_row = work_quality_control(ws, begin_row, data)

    # data_end.update(get_data_created(data_end_row, data["data_of_created"]))

    data_in_cell(ws, data_end)

    # ws.merge_cells(f'I{data_end_row + 3}:J{data_end_row + 3}')
    merge_end(ws, data_end_row - 1)

    return data_end_row


# работа со втором листом
def work_sheet_2(ws, data):
    amount_method = main_work_sh2(ws, data)
    begin_row = amount_method + 13

    end_work_sh2(ws, begin_row, data)


# заполняем список задач пропусками, потом проходим по заданным в petrophysic_selected задачам и задаем соотв значения
def get_petrophysic_selected(petrophysic_selected):
    petrophysics_value = {}
    for petrophysic_name in Inform_for_method.Types_task:
        petrophysics_value.update({petrophysic_name: "-"})
    if len(petrophysic_selected) != 0:
        for petrophysic in petrophysic_selected:
            petrophysics_value[petrophysic] = PETROPHYSICS[petrophysic]
    return petrophysics_value


# формирование dict для списка петрофизических задач для метода
def get_petrofisic_data(numb_row, method):
    cells = [f"A{numb_row}", f"B{numb_row}", f"E{numb_row}", f"G{numb_row}", f"I{numb_row}", f"K{numb_row}",
             f"M{numb_row}"]
    keys = ["method"] + Inform_for_method.Types_task + ["method_value"]

    petrofisic_data = {"method": method["method"]}
    petrofisic_data.update(get_petrophysic_selected(method["petrophysic_selected"]))
    petrofisic_data.update({"method_value": method["method_value"]})
    return get_data_keys(petrofisic_data, keys, cells, italic=True)


# формирование таблицы петрофизических задач для методов
def table_data_sh3(first_row, inform_for_method):
    petrofisic_inf = {}
    numb_row = first_row
    for method in inform_for_method:
        petrofisic_inf.update(get_petrofisic_data(numb_row, method))
        numb_row = numb_row + 1
    return petrofisic_inf


# возвращаем данные по петрофизическим задачам + количество методов
def get_main_data_sh3(data, first_row):
    inform_for_method = data["inform_for_method"]
    amount_method = len(inform_for_method)

    return table_data_sh3(first_row, inform_for_method), amount_method


# слияние клеток для третьего листа
def merge_sh3(ws, amount_method):
    for i in range(1, amount_method):
        row = i + 12
        ws.merge_cells(f'B{row}:D{row}')
        ws.merge_cells(f'E{row}:F{row}')
        ws.merge_cells(f'G{row}:H{row}')
        ws.merge_cells(f'I{row}:J{row}')
        ws.merge_cells(f'K{row}:L{row}')
        ws.merge_cells(f'M{row}:N{row}')


# основная работа с третьим листом - петрофизические задачи
def main_work_sh3(ws, data):
    methods_inform, amount_method = get_main_data_sh3(data, 12)
    ws.move_range("A13:M24", rows=(amount_method - 1), cols=0, translate=True)
    style_row(ws, amount_method)

    data_in_cell(ws, methods_inform)

    main_cell_height(ws, amount_method)
    # set_auto_height(ws, amount_method)

    return amount_method


# работа с информацией внизу 3 листа (quality_control)
def end_work_sh3(ws, begin_row, data):
    data_end, data_end_row = work_quality_control(ws, begin_row, data)

    data_in_cell(ws, data_end)

    return data_end_row


# работа с 3 листом
def work_sheet_3(ws, data):
    amount_method = main_work_sh3(ws, data)
    begin_row = amount_method + 13

    data_end_row = end_work_sh3(ws, begin_row, data)

    # get_auto_height(ws)

    merge_end(ws, data_end_row - 1)
    merge_sh3(ws, amount_method)


# работа с 4 листом
def work_sheet_4(ws):
    pass


# установка типа Border клеткам
def set_border(ws):
    for row in ws.iter_rows(min_row=1, min_col=1, max_col=13, max_row=60):
        for cell in row:
            cell.border = Border(left=Side(border_style="thin", color='FFFFFF'),
                                 right=Side(border_style="thin", color='FFFFFF'),
                                 top=Side(border_style="thin", color='FFFFFF'),
                                 bottom=Side(border_style="thin", color='FFFFFF'))


# распределение работ по листам
def separate_work_ws(wb, data, creater):
    for sheet in wb:
        if sheet.title == SheetNames.SHEET_1.value:
            work_sheet_1(sheet, data, creater)
        elif sheet.title == SheetNames.SHEET_2.value:
            work_sheet_2(sheet, data)
        elif sheet.title == SheetNames.SHEET_3.value:
            work_sheet_3(sheet, data)
        elif sheet.title == SheetNames.SHEET_4.value:
            work_sheet_4(sheet)

        # work_cell_width(sheet)
        set_border(sheet)


# сохранение отчета в xlsx
def save_xlsx(data, pk, creater):
    wb = get_template()
    # sheets = get_sheets(wb)

    ws = wb.get_sheet_by_name("Лист2. Оценка качества ГИС")
    # for row in ws.values:
    #     for value in row:
    #         print(value)
    # print(tuple(ws.rows))

    # general_work_ws(wb, get_general_data(data))  # передавать данные хедера
    # separate_work_ws(wb, data, creater)

    # create_name_file = f'{data["customer_name"]}_{data["field_name"]}_{data["num_pad"]}_{data["num_well"]}_{pk}'
    # name_file = '_'.join(create_name_file.split())
    # wb.save(f'files_root\\checklist\\{name_file}.xlsx')

    wb.close()
    # return name_file


# находим template
def get_template():  # передавать доп параметры в название отчета
    return load_workbook('files_root\\checklist_template_v5.xlsx')


# сохранение отчета в pdf
def save_pdf(data, pk, creater):
    name_file = save_xlsx(data, pk, creater)

    # df = pd.read_excel(f'files_root\\checklist\\{name_file}.xlsx')
    # df.to_html(f'files_root\\checklist\\{name_file}.html')
    # config = pdfkit.configuration(wkhtmltopdf='files_root\\wkhtmltopdf\\bin\\wkhtmltopdf.exe')
    # pdfkit.from_file(f'files_root\\checklist\\{name_file}.html', f'files_root\\checklist\\{name_file}.pdf',
    #                  configuration=config)

    book = xw.Book(f'files_root\\checklist\\{name_file}.xlsx')

    book.to_pdf(f'files_root\\checklist\\{name_file}.pdf')

    app = book.app
    # book.api.ExportAsFixedFormat(0, f'files_root\\checklist\\{name_file}.pdf')
    app.visible = False

    app.kill()

    return name_file