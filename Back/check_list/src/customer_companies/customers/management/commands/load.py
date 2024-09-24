import os
import sys
from openpyxl import load_workbook

from authorization_for_check.service_logics import del_perm
from check_list.settings import BASE_DIR
from customer_companies.clusters.models import Cluster
from customer_companies.customers.models import Customer
from customer_companies.fields.models import Field
from customer_companies.strata.models import Strata
from customer_companies.wellbores.models import Wellbore
from customer_companies.wells.models import Well
from service_companies.services.models import Service


def create(class_name, **kwargs):
    og = getattr(sys.modules[__name__], class_name)(**kwargs)
    og.save()


def create_og(sheet):
    for row in sheet.rows:
        cell = row[1]
        if cell.row > 3 and cell.value:
            try:
                customer = Customer.objects.get(name=cell.value)
                if customer.time_deleted is not None:
                    del_perm(customer)
                    customer.delete()
                    create(class_name="Customer", **{"name": cell.value})
            except:
                create(class_name="Customer", **{"name": cell.value})


def create_field(sheet):
    # for row in range(4, 8):
    for row in sheet.rows:
        cell = row[2]
        if cell.row > 3 and cell.value:
            try:
                customer = Customer.objects.get(name=sheet.cell(column=2, row=cell.row).value)
                field = Field.objects.get(name=cell.value, customer=customer.pk)
                if field.time_deleted is not None:
                    del_perm(field)
                    field.delete()
                    create(class_name="Field", **{"name": cell.value, "customer": customer})
            except Customer.DoesNotExist:
                pass
            except Field.DoesNotExist:
                create(class_name="Field", **{"name": cell.value, "customer": customer})


def create_cluster(sheet):
    for row in sheet.rows:
        cell = row[3]
        if cell.row > 3 and cell.value:
            try:
                field = Field.objects.get(name=sheet.cell(column=3, row=cell.row).value)
                cluster = Cluster.objects.get(name=cell.value, field=field.pk)
                if cluster.time_deleted is not None:
                    del_perm(cluster)
                    cluster.delete()
                    create(class_name="Cluster", **{"name": cell.value, "field": field})
            except Field.DoesNotExist:
                pass
            except Cluster.DoesNotExist:
                create(class_name="Cluster", **{"name": cell.value, "field": field})


def create_well(sheet):
    for row in sheet.rows:
        cluster_c = row[3]
        field_c = row[2]
        well_type_c = row[4]
        num_well_c = row[5]
        if num_well_c.row > 3 and (cell.value for cell in [cluster_c, well_type_c, num_well_c]):
            try:
                field = Field.objects.get(name=field_c.value)
                cluster = Cluster.objects.get(name=cluster_c.value, field=field.pk)
                well = Well.objects.get(num_well=num_well_c.value, cluster=cluster.pk,
                                        well_type=well_type_c.value)
                if well.time_deleted:
                    del_perm(well)
                    well.delete()
                    create(class_name="Well", **{"num_well": num_well_c.value, "cluster": cluster,
                                                 "num_pad": cluster.name, "well_type": well_type_c.value})
            except Field.DoesNotExist or Cluster.DoesNotExist:
                pass
            except Well.DoesNotExist:
                create(class_name="Well", **{"num_well": num_well_c.value, "cluster": cluster,
                                             "num_pad": cluster.name, "well_type": well_type_c.value})
            # except Well.MultipleObjectsReturned:
            #     Well.objects.filter(num_well=num_well_c.value, cluster=cluster.pk,
            #                             well_type=well_type_c.value).delete()
            #     create(class_name="Well", **{"num_well": num_well_c.value, "cluster": cluster,
            #                                  "num_pad": cluster.name, "well_type": well_type_c.value})


def get_data_wellbore(cells):
    data_well = {}
    for key in cells.keys():
        if cells[key].value and cells[key].value != "-":
            data_well.update({key: cells[key].value})

    if "ILWD_LU" in data_well.keys() and data_well["ILWD_LU"]:
        data_well["ILWD_LU"] = "Выполнена"
    else:
        data_well["ILWD_LU"] = "Не выполнена"

    for key in ["ILWD_A", "ILQC_A", "ILQC_C"]:
        if key in data_well.keys() and data_well[key]:
            data_well[key] = " ".join(list(reversed(data_well[key].split())))

    return data_well


def get_wellbore_cells(row):
    return {"pie_well": row[6], "num_wellbore": row[7], "diametr": row[8], "status_wellbore": row[9],
            "ILWD_I": row[10], "main_strata": row[11], "contractor": row[12], "WP_TD": row[13],
            "ILWD_LU": row[14], "WP_DM": row[15], "ILWD_TFS": row[16], "ILWD_TLS": row[17], "ILWD_TRS": row[18],
            "ILWD_A": row[19], "ILWD_TM": row[20], "ILQC_A": row[21], "ILQC_C": row[22], "ILQC_TR": row[23],
            "WP_PCS": row[24], "WP_PT1": row[25], "WP_PT3": row[26], "WP_PWL": row[27], "WP_PCP": row[28],
            "WP_CS": row[29], "WP_T1": row[30], "WP_T3": row[31], "WP_WL": row[32], "WP_CP": row[33],
            "WP_GRemark": row[34]}


def get_create_service(value):
    try:
        service = Service.objects.get(name=value)
    except Service.DoesNotExist:
        if value:
            create(class_name="Service", **{"name": value})


def get_create_strata(value, field):
    try:
        strata = Strata.objects.get(name=value, field=field.pk)
    except Strata.DoesNotExist:
        if value:
            create(class_name="Strata", **{"name": value, "field": field})


def create_wellbore(sheet):
    for row in sheet.rows:
        cluster_c = row[3]
        field_c = row[2]
        well_type_c = row[4]
        num_well_c = row[5]

        if num_well_c.row > 3 and cluster_c.value and field_c.value and well_type_c.value and num_well_c.value \
                and row[6].value and row[7].value:
            cells = get_wellbore_cells(row)
            data_well = get_data_wellbore(cells)
            try:
                field = Field.objects.get(name=field_c.value)
                cluster = Cluster.objects.get(name=cluster_c.value, field=field.pk)
                well = Well.objects.get(num_well=num_well_c.value, cluster=cluster.pk, well_type=well_type_c.value)

                get_create_service(cells["contractor"].value)
                get_create_strata(cells["main_strata"].value, field)

                wellbore = Wellbore.objects.get(num_wellbore=cells["num_wellbore"].value,
                                                pie_well=cells["pie_well"].value, well=well.pk)

                if wellbore.time_deleted:
                    del_perm(wellbore)
                    wellbore.delete()
                    data_well.update({"well": well})
                    create(class_name="Wellbore", **data_well)
            except Cluster.DoesNotExist or Field.DoesNotExist or Well.DoesNotExist:
                pass
            except Wellbore.DoesNotExist:
                data_well.update({"well": well})
                create(class_name="Wellbore", **data_well)
            # except Wellbore.MultipleObjectsReturned:
            #     Wellbore.objects.filter(num_wellbore=cells["num_wellbore"].value,
            #                                     pie_well=cells["pie_well"].value, well=well.pk).delete()
            #     data_well.update({"well": well})
            #     create(class_name="Wellbore", **data_well)


def load_database():
    book = load_workbook(os.path.join(BASE_DIR, "load.xlsx"))

    sheet = book[book.sheetnames[0]]

    # create_og(sheet)

    # create_field(sheet)
    #
    # create_cluster(sheet)
    #
    # create_well(sheet)
    #
    create_wellbore(sheet)
